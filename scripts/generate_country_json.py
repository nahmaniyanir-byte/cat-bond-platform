#!/usr/bin/env python3
"""
Generate country JSON datasets from Excel workbooks under data/countries/{country}/.

Usage:
  python scripts/generate_country_json.py
"""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
COUNTRIES_ROOT = PROJECT_ROOT / "data" / "countries"
INDEX_PATH = COUNTRIES_ROOT / "country_page_index.json"
GLOBE_POINTS_PATH = COUNTRIES_ROOT / "country_globe_points_country_files.json"

FILE_KEY_MAP = {
    "COUNTRYOVERVIEW": ("country_overview", "country_overview.json"),
    "DISASTER": ("disaster_context", "disaster_context.json"),
    "DOCUMENTMETADATA": ("document_metadata", "document_metadata.json"),
    "DOCUMENMETADATA": ("document_metadata", "document_metadata.json"),
    "INVESTORDISTRIBUTION": ("investor_distribution", "investor_distribution.json"),
    "INVESTORGEOGRAPHY": ("investor_geography", "investor_geography.json"),
    "POLICYLESSONS": ("policy_lessons", "policy_lessons.json"),
    "POLICYLESSONSX": ("policy_lessons", "policy_lessons.json"),
    "PRICINGDETAILS": ("pricing_details", "pricing_details.json"),
    "RISKPARAMETERS": ("risk_parameters", "risk_parameters.json"),
    "STRUCTURECOVERAGE": ("structure_coverage", "structure_coverage.json"),
    "TRANSACTIONDETAILS": ("transaction_details", "transaction_details.json"),
}

SECTION_JSON_FILE_MAP = {
    "country_overview": "country_overview.json",
    "disaster_context": "disaster_context.json",
    "document_metadata": "document_metadata.json",
    "investor_distribution": "investor_distribution.json",
    "investor_geography": "investor_geography.json",
    "policy_lessons": "policy_lessons.json",
    "pricing_details": "pricing_details.json",
    "risk_parameters": "risk_parameters.json",
    "structure_coverage": "structure_coverage.json",
    "transaction_details": "transaction_details.json",
}

EXPECTED_SECTION_KEYS = [
    "country_overview",
    "disaster_context",
    "document_metadata",
    "investor_distribution",
    "investor_geography",
    "policy_lessons",
    "pricing_details",
    "risk_parameters",
    "structure_coverage",
    "transaction_details",
]

COUNTRY_COORDS = {
    "mexico": {"lat": 23.6345, "lng": -102.5528},
    "chile": {"lat": -35.6751, "lng": -71.5430},
    "colombia": {"lat": 4.5709, "lng": -74.2973},
    "peru": {"lat": -9.1900, "lng": -75.0152},
    "philippines": {"lat": 12.8797, "lng": 121.7740},
    "jamaica": {"lat": 18.1096, "lng": -77.2975},
    "israel": {"lat": 31.0461, "lng": 34.8516},
}

COUNTRY_REGIONS = {
    "mexico": "Latin America",
    "chile": "Latin America",
    "colombia": "Latin America",
    "peru": "Latin America",
    "philippines": "Asia",
    "jamaica": "Caribbean",
    "israel": "Middle East",
}


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower().strip())
    return re.sub(r"(^-+|-+$)", "", slug)


def normalize_filename_key(file_name: str) -> str:
    stem = Path(file_name).stem.upper()
    return re.sub(r"[^A-Z0-9]", "", stem)


def normalize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        cleaned = value.strip().replace("\u00a0", " ")
        return cleaned if cleaned else None
    return value


def is_empty_row(row: dict[str, Any]) -> bool:
    return not any(value is not None and value != "" for value in row.values())


def read_sheet_rows(xlsx_path: Path) -> tuple[str, list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    raw_rows = list(sheet.iter_rows(values_only=True))
    header_idx = None
    for idx, row in enumerate(raw_rows):
        if any(cell is not None and str(cell).strip() for cell in row):
            header_idx = idx
            break

    if header_idx is None:
        return workbook.sheetnames[0], []

    headers = [str(cell).strip() if cell is not None else f"column_{i+1}" for i, cell in enumerate(raw_rows[header_idx])]
    records: list[dict[str, Any]] = []

    for row in raw_rows[header_idx + 1 :]:
        record = {headers[i]: normalize_value(row[i]) if i < len(row) else None for i in range(len(headers))}
        if not is_empty_row(record):
            records.append(record)

    return workbook.sheetnames[0], records


def parse_money(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return 0.0
    text = str(value).strip().lower().replace(",", "")
    if not text:
        return 0.0

    # examples: "$500m", "$1.2bn", "100000000"
    text = text.replace("$", "")
    mult = 1.0
    if text.endswith("bn") or text.endswith("b"):
        mult = 1_000_000_000
        text = re.sub(r"(bn|b)$", "", text)
    elif text.endswith("mn") or text.endswith("m"):
        mult = 1_000_000
        text = re.sub(r"(mn|m)$", "", text)

    try:
        return float(text) * mult
    except ValueError:
        return 0.0


def parse_year(value: Any) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        if re.fullmatch(r"\d{4}", value.strip()):
            return int(value.strip())
        date_match = re.search(r"\b(19|20)\d{2}\b", value)
        if date_match:
            return int(date_match.group(0))
    return None


def infer_sovereign_flag(rows_by_section: dict[str, list[dict[str, Any]]]) -> bool:
    overview = rows_by_section.get("country_overview", [])
    for row in overview:
        sponsor_type = str(row.get("Sponsor Type", "") or "").lower()
        sponsor = str(row.get("Sponsor", "") or "").lower()
        if "sovereign" in sponsor_type:
            return True
        if any(token in sponsor for token in ["government", "republic", "ministry", "treasury", "fonden"]):
            return True

    transactions = rows_by_section.get("transaction_details", [])
    for row in transactions:
        protection = str(row.get("Protection Type", "") or "").lower()
        if "sovereign" in protection:
            return True
    return False


def derive_country_kpis(country_name: str, slug: str, rows_by_section: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    overview = rows_by_section.get("country_overview", [])
    transactions = rows_by_section.get("transaction_details", [])
    disaster = rows_by_section.get("disaster_context", [])

    deal_count = len(transactions) if transactions else len(overview)

    total_volume_usd = 0.0
    if transactions:
        for row in transactions:
            total_volume_usd += parse_money(row.get("Size USD"))
    elif overview:
        for row in overview:
            total_volume_usd += parse_money(row.get("Deal Size"))

    years = []
    for row in transactions:
        issue_date = row.get("Issue Date")
        if isinstance(issue_date, str):
            y = parse_year(issue_date)
            if y:
                years.append(y)
    for row in overview:
        y = parse_year(row.get("Year"))
        if y:
            years.append(y)

    latest_issue_year = max(years) if years else None

    peril_candidates = []
    trigger_candidates = []
    for row in overview:
        peril_val = row.get("Peril")
        if peril_val:
            peril_candidates.append(str(peril_val))
        trigger_val = row.get("Trigger Type")
        if trigger_val:
            trigger_candidates.append(str(trigger_val))
    for row in disaster:
        hazard = row.get("Main Hazard")
        if hazard:
            peril_candidates.append(str(hazard))

    main_peril = Counter(peril_candidates).most_common(1)[0][0] if peril_candidates else "Not stated"
    main_trigger_type = Counter(trigger_candidates).most_common(1)[0][0] if trigger_candidates else "Not stated"

    sovereign_flag = infer_sovereign_flag(rows_by_section)
    market_segment = "Sovereign" if sovereign_flag else "Non-Sovereign / Corporate"

    return {
        "country_name": country_name,
        "slug": slug,
        "total_deals": deal_count,
        "total_volume_usd": round(total_volume_usd),
        "latest_issue_year": latest_issue_year,
        "main_peril": main_peril,
        "main_trigger_type": main_trigger_type,
        "sovereign_flag": sovereign_flag,
        "market_segment": market_segment,
    }


def derive_timeline(rows_by_section: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    timeline: list[dict[str, Any]] = []
    transactions = rows_by_section.get("transaction_details", [])
    overview_map = {str(row.get("Deal Name")): row for row in rows_by_section.get("country_overview", [])}

    for row in transactions:
        deal_name = row.get("Deal Name")
        issue_date = row.get("Issue Date")
        year = parse_year(issue_date) if issue_date else None
        overview_row = overview_map.get(str(deal_name))
        if not year and overview_row:
            year = parse_year(overview_row.get("Year"))

        timeline.append(
            {
                "deal_name": deal_name,
                "issue_date": issue_date,
                "year": year,
                "size_usd": round(parse_money(row.get("Size USD"))),
                "covered_region": row.get("Covered Region"),
                "protection_type": row.get("Protection Type"),
                "coverage_basis": row.get("Coverage Basis"),
            }
        )

    timeline.sort(key=lambda item: (item.get("year") or 0, str(item.get("issue_date") or "")))
    return timeline


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def build_country_summary(
    country_name: str,
    slug: str,
    folder_name: str,
    kpis: dict[str, Any],
    rows_by_section: dict[str, list[dict[str, Any]]],
    timeline: list[dict[str, Any]],
) -> dict[str, Any]:
    overview = rows_by_section.get("country_overview", [])
    summary_text = overview[0].get("Short Summary") if overview else None
    available_sections = [section for section in EXPECTED_SECTION_KEYS if rows_by_section.get(section)]

    trigger_values = sorted(
        {
            str(row.get("Trigger Type")).strip()
            for row in overview
            if row.get("Trigger Type") and str(row.get("Trigger Type")).strip()
        }
    )
    model_types = sorted(
        {
            str(row.get("Sponsor Type")).strip()
            for row in overview
            if row.get("Sponsor Type") and str(row.get("Sponsor Type")).strip()
        }
    )

    region = COUNTRY_REGIONS.get(slug) or str(overview[0].get("Region") if overview else "Not stated")
    if not region or region.lower().startswith("not stated"):
        region = COUNTRY_REGIONS.get(slug, "Not stated")

    return {
        "country_name": country_name,
        "slug": slug,
        "folder_name": folder_name,
        "region": region,
        "summary": summary_text or "Country intelligence profile generated from structured Excel datasets.",
        "available_sections": available_sections,
        "has_investor_data": bool(rows_by_section.get("investor_distribution")),
        "has_investor_geography": bool(rows_by_section.get("investor_geography")),
        "has_policy_lessons": bool(rows_by_section.get("policy_lessons")),
        "has_documents": bool(rows_by_section.get("document_metadata")),
        "has_transactions": bool(rows_by_section.get("transaction_details")),
        "deal_count": kpis["total_deals"],
        "total_deals": kpis["total_deals"],
        "total_volume_usd": kpis["total_volume_usd"],
        "total_volume": kpis["total_volume_usd"],
        "total_volume_musd": round(kpis["total_volume_usd"] / 1_000_000, 2),
        "latest_issue_year": kpis["latest_issue_year"],
        "main_peril": kpis["main_peril"],
        "trigger_types": trigger_values or [kpis["main_trigger_type"]],
        "model_types": model_types or [kpis["market_segment"]],
        "sovereign_flag": kpis["sovereign_flag"],
        "market_segment_summary": kpis["market_segment"],
        "destination_url": f"/countries/{slug}",
        "timeline_count": len(timeline),
    }


def build_globe_point(summary: dict[str, Any]) -> dict[str, Any] | None:
    slug = summary["slug"]
    coords = COUNTRY_COORDS.get(slug)
    if not coords:
        return None
    tooltip = (
        f'{summary["deal_count"]} deal(s) | {summary["market_segment_summary"]} | '
        f'Latest: {summary["latest_issue_year"] or "N/A"}'
    )
    return {
        "id": slug,
        "country_name": summary["country_name"],
        "slug": slug,
        "lat": coords["lat"],
        "lng": coords["lng"],
        "sovereign_flag": bool(summary["sovereign_flag"]),
        "market_segment": summary["market_segment_summary"],
        "deal_count": summary["deal_count"],
        "total_volume_usd": summary["total_volume_usd"],
        "main_peril": summary["main_peril"],
        "latest_issue_year": summary["latest_issue_year"],
        "tooltip_title": summary["country_name"],
        "tooltip_text": tooltip,
        "destination_url": f"/countries/{slug}",
    }


def process_country(country_dir: Path) -> dict[str, Any] | None:
    excel_files = sorted(country_dir.glob("*.xlsx"), key=lambda p: p.name.lower())
    if not excel_files:
        return None

    rows_by_section: dict[str, list[dict[str, Any]]] = {key: [] for key in EXPECTED_SECTION_KEYS}
    source_map: dict[str, str] = {}
    country_name: str | None = None
    slug = slugify(country_dir.name)

    for xlsx_path in excel_files:
        file_key = normalize_filename_key(xlsx_path.name)
        mapped = FILE_KEY_MAP.get(file_key)
        if not mapped:
            continue

        section_key, json_file_name = mapped
        sheet_name, rows = read_sheet_rows(xlsx_path)
        rows_by_section[section_key] = rows
        source_map[section_key] = xlsx_path.name

        if rows and not country_name:
            maybe_country = rows[0].get("Country")
            if maybe_country:
                country_name = str(maybe_country)

        out_path = country_dir / "json" / json_file_name
        write_json(
            out_path,
            {
                "country": country_name or country_dir.name,
                "slug": slug,
                "source_file": xlsx_path.name,
                "sheet_name": sheet_name,
                "row_count": len(rows),
                "rows": rows,
            },
        )

    for section_key in EXPECTED_SECTION_KEYS:
        if section_key in source_map:
            continue
        out_path = country_dir / "json" / SECTION_JSON_FILE_MAP[section_key]
        write_json(
            out_path,
            {
                "country": country_name or country_dir.name,
                "slug": slug,
                "source_file": None,
                "sheet_name": None,
                "row_count": 0,
                "rows": [],
            },
        )

    country_name = country_name or country_dir.name
    kpis = derive_country_kpis(country_name, slug, rows_by_section)
    timeline = derive_timeline(rows_by_section)
    summary = build_country_summary(country_name, slug, country_dir.name, kpis, rows_by_section, timeline)

    write_json(country_dir / "json" / "country_kpis.json", kpis)
    write_json(
        country_dir / "json" / "timeline.json",
        {"country": country_name, "slug": slug, "count": len(timeline), "events": timeline},
    )
    write_json(
        country_dir / "json" / "country_page_summary.json",
        {
            "country": country_name,
            "slug": slug,
            "summary": summary,
            "source_files": source_map,
        },
    )

    return summary


def main() -> None:
    country_summaries: list[dict[str, Any]] = []
    globe_points: list[dict[str, Any]] = []

    for country_dir in sorted([p for p in COUNTRIES_ROOT.iterdir() if p.is_dir()], key=lambda p: p.name.lower()):
        summary = process_country(country_dir)
        if not summary:
            continue
        country_summaries.append(summary)
        point = build_globe_point(summary)
        if point:
            globe_points.append(point)

    country_summaries.sort(key=lambda item: item["country_name"].lower())
    globe_points.sort(key=lambda item: item["country_name"].lower())

    write_json(
        INDEX_PATH,
        {
            "generated_at": datetime.now(UTC).isoformat(),
            "source_root": str(COUNTRIES_ROOT),
            "countries": country_summaries,
        },
    )
    write_json(GLOBE_POINTS_PATH, globe_points)

    print(f"Generated country JSON for {len(country_summaries)} countries.")
    print(f"Index written: {INDEX_PATH}")
    print(f"Globe points written: {GLOBE_POINTS_PATH}")


if __name__ == "__main__":
    main()
